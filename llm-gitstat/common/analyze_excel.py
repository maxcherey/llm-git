import datetime
import json
import logging
import os
import re
from typing import OrderedDict
import openpyxl

from common.stat_contributors import generate_contributors_report
from common.utils import parse_url

ORIGIN_SHEET_NAME="OSS projects"
ORIGIN_VENDOR_COLUMN = 'B'
ORIGIN_VENDOR_WEBSITE_COLUMN = 'C'
ORIGIN_VENDOR_VALUATION_COLUMN = 'D'
ORIGIN_VENDOR_REVENUE_COLUMN = 'E'
ORIGIN_PROJECT_NAME_COLUMN = 'F'
ORIGIN_PROJECT_START_COLUMN = 'G'
ORIGIN_PROJECT_START_YEARS_COLUMN = 'Q'
ORIGIN_PROJECT_CATEGORY_COLUMN = 'H'
ORIGIN_PROJECT_DESCRIPTION_COLUMN = 'I'
ORIGIN_LICENSE_FAMILY_COLUMN = 'P'
ORIGIN_CNCF_COLUMN = 'O'

ORIGIN_KLOC_COLUMN = 'J'
ORIGIN_CONTRIBUTORS_COLUMN = 'K'
ORIGIN_STARS_COLUMN = 'L'
ORIGIN_GITHUB_LINK_COLUMN = 'M'
ORIGIN_LICENSE_COLUMN = 'N'
ORIGIN_PROJECT_START_DATE_COLUMN = 'G'

SCRIPT_SHEET_NAME="OSS projects - auto"
SCRIPT_VENDOR_COLUMN = 'SCRIPT_VENDOR_COLUMN'
SCRIPT_VENDOR_WEBSITE_COLUMN = 'SCRIPT_VENDORS_WEBSITE_COLUMN'
SCRIPT_VENDOR_VALUATION_COLUMN = 'SCRIPT_VENDOR_VALUATION_COLUMN'
SCRIPT_VENDOR_REVENUE_COLUMN = 'SCRIPT_VENDOR_REVENUE_COLUMN'
SCRIPT_PROJECT_NAME_COLUMN = 'SCRIPT_PROJECT_NAME_COLUMN'
SCRIPT_PROJECT_START_ORIGIN_COLUMN = 'SCRIPT_PROJECT_START_ORIGIN_COLUMN'
SCRIPT_PROJECT_START_COLUMN = 'SCRIPT_PROJECT_START_COLUMN'
SCRIPT_PROJECT_START_YEARS_COLUMN = 'SCRIPT_PROJECT_START_YEARS_COLUMN'
SCRIPT_PROJECT_ACTIVE_SCORE_COLUMN = 'SCRIPT_PROJECT_ACTIVE_SCORE_COLUMN'
SCRIPT_PROJECT_CATEGORY_COLUMN = 'SCRIPT_PROJECT_CATEGORY_COLUMN'
SCRIPT_PROJECT_DESCRIPTION_ORIGIN_COLUMN = 'SCRIPT_PROJECT_DESCRIPTION_ORIGIN_COLUMN'
SCRIPT_PROJECT_DESCRIPTION_COLUMN = 'SCRIPT_PROJECT_DESCRIPTION_COLUMN'
SCRIPT_PROJECT_GITHUB_COLUMN = 'SCRIPT_PROJECT_GITHUB_COLUMN'
SCRIPT_STARS_ORIGIN_COLUMN = 'SCRIPT_STARS_ORIGIN_COLUMN'
SCRIPT_STARS_COLUMN = 'SCRIPT_STARS_COLUMN'
SCRIPT_FORKS_COLUMN = 'SCRIPT_FORKS_COLUMN'
SCRIPT_LANGUAGE_COLUMN = 'SCRIPT_LANGUAGE_COLUMN'
SCRIPT_TOP3_FILES_COLUMN = 'SCRIPT_TOP3_FILES_COLUMN'
SCRIPT_KLOC_ORIGIN_COLUMN = 'SCRIPT_KLOC_ORIGIN_COLUMN'
SCRIPT_KLOC_COLUMN = 'SCRIPT_KLOC_COLUMN'
SCRIPT_SIZE_COLUMN = 'SCRIPT_SIZE_COLUMN'
SCRIPT_CONTRIBUTORS_TOTAL_ORIGIN_COLUMN = 'SCRIPT_CONTRIBUTORS_TOTAL_ORIGIN_COLUMN'
SCRIPT_CONTRIBUTORS_TOTAL_COLUMN = 'SCRIPT_CONTRIBUTORS_TOTAL_COLUMN'
SCRIPT_CONTRIBUTORS_TOTAL_CONTRIBUTIONS_COLUMN = 'SCRIPT_CONTRIBUTORS_TOTAL_CONTRIBUTIONS_COLUMN'
SCRIPT_CONTRIBUTORS_TOP3_COLUMN = 'SCRIPT_CONTRIBUTORS_TOP3_COLUMN'
SCRIPT_CONTRIBUTORS_USERS_TOTAL_COLUMN = 'SCRIPT_CONTRIBUTORS_USERS_TOTAL_COLUMN'
SCRIPT_CONTRIBUTORS_USERS_TOTAL_PERCENT_COLUMN = 'SCRIPT_CONTRIBUTORS_USERS_TOTAL_PERCENT_COLUMN'
SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN = 'SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN'
SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN = 'SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN'
SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_COLUMN = 'SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_COLUMN'
SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN = 'SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN'
SCRIPT_CONTRIBUTORS_ANON_TOTAL_COLUMN = 'SCRIPT_CONTRIBUTORS_ANON_TOTAL_COLUMN'
SCRIPT_CONTRIBUTORS_ANON_TOTAL_PERCENT_COLUMN = 'SCRIPT_CONTRIBUTORS_ANON_TOTAL_PERCENT_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_PERCENT_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_PERCENT_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_P_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_P_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_COLUMN'
SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_P_COLUMN = 'SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_P_COLUMN'
SCRIPT_LICENSE_ORIGIN_COLUMN = 'SCRIPT_LICENSE_ORIGIN_COLUMN'
SCRIPT_LICENSE_COLUMN = 'SCRIPT_LICENSE_COLUMN'
SCRIPT_LICENSE_FAMILY_COLUMN = 'SCRIPT_LICENSE_FAMILY_COLUMN'
SCRIPT_CNCF_COLUMN = 'SCRIPT_CNCF_COLUMN'
SCRIPT_HAS_COMMERCIAL_COMPANY_COLUMN = 'SCRIPT_HAS_COMMERCIAL_COMPANY_COLUMN'
SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_COLUMN = 'SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_COLUMN'
SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_WITH_ORG_COLUMN = 'SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_WITH_ORG_COLUMN'
SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_COLUMN = 'SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_COLUMN'
SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_WITHIN_ORG_COLUMN = 'SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_WITHIN_ORG_COLUMN'
SCRIPT_TOP10_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN = 'SCRIPT_TOP10_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN'
SCRIPT_INTERNAL_DOMAINS_USED_COLUMN = 'SCRIPT_INTERNAL_DOMAINS_USED_COLUMN'


SCRIPT_SHEET_COLUMN_NAMES = OrderedDict([
    (SCRIPT_VENDOR_COLUMN, 'Vendor'),
    (SCRIPT_VENDOR_WEBSITE_COLUMN, 'Vendor Website'),
    (SCRIPT_VENDOR_VALUATION_COLUMN,'Vendor Valuation, 2023($M)'),
    (SCRIPT_VENDOR_REVENUE_COLUMN, 'Vendor Revenue, 2023($M)'),
    (SCRIPT_PROJECT_NAME_COLUMN,'Project Name'),
    (SCRIPT_PROJECT_START_ORIGIN_COLUMN, 'Project Start (ChatGPT)'),
    (SCRIPT_PROJECT_START_COLUMN, 'Project Start (GitHub)'),
    (SCRIPT_PROJECT_START_YEARS_COLUMN, 'Project start, years (ChatGPT)'),
    (SCRIPT_PROJECT_ACTIVE_SCORE_COLUMN, 'Project Active Score)'),
    (SCRIPT_PROJECT_CATEGORY_COLUMN,'Project Category'),
    (SCRIPT_PROJECT_DESCRIPTION_ORIGIN_COLUMN, 'Project Description (ChatGPT)'),
    (SCRIPT_PROJECT_DESCRIPTION_COLUMN, 'Project Description (GitHub)'),
    (SCRIPT_PROJECT_GITHUB_COLUMN, 'GitHub Link'),
    (SCRIPT_STARS_ORIGIN_COLUMN, 'Stars (ChatGPT)'),
    (SCRIPT_STARS_COLUMN, 'Stars (GitHub)'),
    (SCRIPT_FORKS_COLUMN,'Forks (GitHub)'),
    (SCRIPT_LANGUAGE_COLUMN, 'Language'),
    (SCRIPT_TOP3_FILES_COLUMN, 'Top3 files'),
    (SCRIPT_KLOC_ORIGIN_COLUMN, 'KLOC (ChatGPT)'),
    (SCRIPT_KLOC_COLUMN, 'KLOC (GitHub)'),
    (SCRIPT_SIZE_COLUMN, 'Size MB (GitHub)'),
    (SCRIPT_CONTRIBUTORS_TOTAL_ORIGIN_COLUMN, 'Total contributors(ChatGPT)'),
    (SCRIPT_CONTRIBUTORS_TOTAL_COLUMN, 'Total contributors (GitHub)'),
    (SCRIPT_CONTRIBUTORS_USERS_TOTAL_COLUMN, 'Total contributors: github users (github)'),
    (SCRIPT_CONTRIBUTORS_USERS_TOTAL_PERCENT_COLUMN, 'Total contributors: github users % (github)'),
    (SCRIPT_CONTRIBUTORS_ANON_TOTAL_COLUMN, 'Contributors Anon (github)'),
    (SCRIPT_CONTRIBUTORS_ANON_TOTAL_PERCENT_COLUMN, 'Contributors Anon % (github)'),
    (SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN, 'Internal contributors (GitHub)'),
    (SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN, 'Internal contributors %'),
    (SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_COLUMN, 'Others contributor'),
    (SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN, 'Other contributor %'),
    (SCRIPT_CONTRIBUTORS_TOTAL_CONTRIBUTIONS_COLUMN, 'Total contributions (github)'),
    (SCRIPT_CONTRIBUTORS_TOP3_COLUMN, 'Top3 contributors (github)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_COLUMN, 'Total contributors for last year (git commits)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_PERCENT_COLUMN, 'Total contributors for last year % (git commits)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN, 'Internal contributors for last year (git commits)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN, 'Internal contributors for last year % (git commits)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_COLUMN, 'Others contributor for last year (git commits)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN, 'Others contributor for last year % (git commits)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN, 'Contributors Top10 domains (git commits)'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_COLUMN, 'Contributors presented on GitHub'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_P_COLUMN, 'Contributors presented on GitHub %'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_COLUMN, 'Contributors as GitHub users'),
    (SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_P_COLUMN, 'Contributors as GitHub users %'),
    (SCRIPT_LICENSE_ORIGIN_COLUMN, 'License'),
    (SCRIPT_LICENSE_COLUMN, 'License (github)'),
    (SCRIPT_LICENSE_FAMILY_COLUMN, 'License Family'),
    (SCRIPT_CNCF_COLUMN, 'Is project part of CNCF'),
    (SCRIPT_HAS_COMMERCIAL_COMPANY_COLUMN, 'Has commercial company'),
    (SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_COLUMN, 'Contributors top10 domains (GitHub)'),
    (SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_WITH_ORG_COLUMN, 'Contributors top10 domains within org (GitHub)'),
    (SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_COLUMN, 'Contributors top10 companies (GitHub)'),
    (SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_WITHIN_ORG_COLUMN, 'Contributors top10 companies within org (GitHub)'),
    (SCRIPT_TOP10_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN, 'Active contributors top10 domains'),
    (SCRIPT_INTERNAL_DOMAINS_USED_COLUMN, 'Used internal domains'),
    ])

SCRIPT_COLUMNS_MAP={}


LIST_TO_UPDATE_GIT = []
LIST_TO_UPDATE = []

def column_number_to_letter(column_number):
  if column_number < 1:
    raise ValueError("Column number must be positive.")

  column_letter = ""
  while column_number > 0:
    remainder = column_number % 26
    if remainder == 0:
      column_letter = "Z" + column_letter
      column_number -= 26
    else:
      column_letter = chr(ord('A') + remainder - 1) + column_letter
      column_number //= 26

  return column_letter


class ExcelEntry: 
    def __init__(self, sheet, script_sheet, index, id, url):
        self.sheet = sheet
        self.script_sheet = script_sheet
        self.row_index = index
        self.id = id
        self.url = url

    def copy_custom_values(self):
        self.set_cell_value(SCRIPT_VENDOR_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_VENDOR_COLUMN)).value)
        self.set_cell_value(SCRIPT_VENDOR_WEBSITE_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_VENDOR_WEBSITE_COLUMN)).value)
        self.set_cell_value(SCRIPT_VENDOR_VALUATION_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_VENDOR_VALUATION_COLUMN)).value)
        self.set_cell_value(SCRIPT_VENDOR_REVENUE_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_VENDOR_REVENUE_COLUMN)).value)
        self.set_cell_value(SCRIPT_PROJECT_NAME_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_PROJECT_NAME_COLUMN)).value)
        self.set_cell_value(SCRIPT_PROJECT_START_ORIGIN_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_PROJECT_START_COLUMN)).value)
        self.set_cell_value(SCRIPT_PROJECT_START_YEARS_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_PROJECT_START_YEARS_COLUMN)).value)
        self.set_cell_value(SCRIPT_PROJECT_CATEGORY_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_PROJECT_CATEGORY_COLUMN)).value)
        self.set_cell_value(SCRIPT_PROJECT_DESCRIPTION_ORIGIN_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_PROJECT_DESCRIPTION_COLUMN)).value)
        self.set_cell_link(SCRIPT_PROJECT_GITHUB_COLUMN, "GitHub", self.url)
        self.set_license_family(SCRIPT_LICENSE_FAMILY_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_LICENSE_FAMILY_COLUMN)).value)
        self.set_cell_value(SCRIPT_CNCF_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_CNCF_COLUMN)).value)
        self.set_cell_value(SCRIPT_KLOC_ORIGIN_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_KLOC_COLUMN)).value)
        self.set_cell_value(SCRIPT_LICENSE_ORIGIN_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_LICENSE_COLUMN)).value)
        self.set_cell_value(SCRIPT_STARS_ORIGIN_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_STARS_COLUMN)).value)
        self.set_cell_value(SCRIPT_CONTRIBUTORS_TOTAL_ORIGIN_COLUMN, self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_CONTRIBUTORS_COLUMN)).value)
        self.set_cell_value(SCRIPT_HAS_COMMERCIAL_COMPANY_COLUMN, 0)


    def get_vendor(self):
        return self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_VENDOR_COLUMN)).value

    def get_license(self):
        return self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_LICENSE_COLUMN)).value
    
    def get_start_year(self):
        return self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_PROJECT_START_DATE_COLUMN)).value

    def get_kloc(self):
        return self.sheet.cell(row=self.row_index, column=openpyxl.utils.column_index_from_string(ORIGIN_KLOC_COLUMN)).value
    
    def set_cell_value(self, column, value):
        self.script_sheet.cell(row=self.row_index, column=SCRIPT_COLUMNS_MAP[column]).value = value

    def set_cell_link(self, column, value, link):
        self.script_sheet.cell(row=self.row_index, column=SCRIPT_COLUMNS_MAP[column]).value = value
        self.script_sheet.cell(row=self.row_index, column=SCRIPT_COLUMNS_MAP[column]).hyperlink = link

    def set_license_family(self, column, value):
        if value is not None and value != "":
            value = value.replace("N", column_number_to_letter(SCRIPT_COLUMNS_MAP[SCRIPT_LICENSE_COLUMN]))
        self.script_sheet.cell(row=self.row_index, column=SCRIPT_COLUMNS_MAP[column]).value = value

def extract_urls(excel_file):
    logging.info(f"Extracting urls from excel file: {excel_file}...")
    workbook = openpyxl.load_workbook(excel_file)
    origin_sheet = workbook.get_sheet_by_name(ORIGIN_SHEET_NAME)
    urls = []
    i = 1
    for row in origin_sheet.iter_rows(values_only=True):
        cell = origin_sheet[f"{ORIGIN_GITHUB_LINK_COLUMN}{i}"]
        if cell.hyperlink:
            urls.append(cell.hyperlink.target)
        i = i + 1
    return urls

def extract_text_urls(text_file):
    logging.info(f"Extracting urls from text file: {text_file}...")
    with open(text_file, 'r') as f:
        urls = f.readlines()
    return [url.strip() for url in urls]

def generate_exec_report(repos, excel_file):
    logging.info(f"Analyzing Excel file: {excel_file}")
    workbook = openpyxl.load_workbook(excel_file)

    # Select the worksheet
    origin_sheet = workbook.get_sheet_by_name(ORIGIN_SHEET_NAME)
    ensure_sheet(workbook)
    script_sheet = workbook.get_sheet_by_name(SCRIPT_SHEET_NAME)
    column = 1
    for key, value in SCRIPT_SHEET_COLUMN_NAMES.items():
        script_sheet.cell(row=1, column=column).value = value
        SCRIPT_COLUMNS_MAP[key] = column
        column += 1

    # Iterate over rows
    data = {}
    i = 2
    for row in origin_sheet.iter_rows(values_only=True):
        cell = origin_sheet[f"M{i}"]
        repo_id = origin_sheet.cell(row=i, column=openpyxl.utils.column_index_from_string(ORIGIN_PROJECT_NAME_COLUMN)).value
        url = ""
        if cell.hyperlink:
            repo_name, owner = parse_url(cell.hyperlink.target)
            if owner == "":
                logging.warning(f"row {i}, vendor: {origin_sheet[f'b{i}'].value}, name: {origin_sheet[f'F{i}'].value}, link {cell.hyperlink.target}")
            repo_id = f"{owner}_{repo_name}"
            url = cell.hyperlink.target
        data[i] = ExcelEntry(origin_sheet, script_sheet, i, repo_id, url)
        i = i + 1
    
    total = 0
    skipped = 0
    for id in data:
        d = data[id]
        d.copy_custom_values()
        total += 1
        if d.url  in repos:
            try:
                repo_data = repos[d.url]
                generate_report(workbook, repo_data, d)
            except ValueError as e:
                logging.warning(f"ValueError generating report for {d.url}: {e}")
                skipped += 1
                if str(e) == "Active days greater than days in period":
                    LIST_TO_UPDATE_GIT.append(d.url)
                continue
            except Exception as e:
                logging.warning(f"Error generating report for {d.url}: {e}")
                skipped += 1
                LIST_TO_UPDATE.append(d.url)
                continue
        workbook.save('modified_oss_projects.xlsx')
    
    logging.info(f"Total: {total}, Skipped: {skipped}")
    if len(LIST_TO_UPDATE) > 0:
        logging.warning(f"List of urls to fix {len(LIST_TO_UPDATE)}:")
        logging.warning(f"{"\n".join(LIST_TO_UPDATE)}")

    if len(LIST_TO_UPDATE_GIT) > 0:
        logging.warning(f"List of urls to fix (GIT) {len(LIST_TO_UPDATE_GIT)}:")
        logging.warning(f"{"\n".join(LIST_TO_UPDATE_GIT)}")

    if len(DOMAINS_TO_FIX) > 0:
        logging.warning(f"List of domains to fix {len(DOMAINS_TO_FIX)}:")
        logging.warning(f"{"\n".join(DOMAINS_TO_FIX)}")


    return data

def ensure_sheet(workbook):
    if SCRIPT_SHEET_NAME not in workbook.sheetnames:
        logging.info(f"Creating '{SCRIPT_SHEET_NAME}' sheet")
        workbook.create_sheet(SCRIPT_SHEET_NAME)
    else:
        logging.info(f"'{SCRIPT_SHEET_NAME}' sheet already exists")
    

def generate_report(workbook, repo, entry):

    #repo created_at
    formatted_date = ""
    if repo.created_at != "":
        datetime_obj = datetime.datetime.fromisoformat(repo.created_at)
        formatted_date = datetime_obj.strftime("%Y")#-%m-%d")
    entry.set_cell_value(SCRIPT_PROJECT_START_COLUMN, f"{formatted_date}")


    logging.info(f"[{entry.row_index}] Updating values for '{entry.sheet.title}', VENDOR: {entry.get_vendor()}: {entry.url}")
    #License
    entry.set_cell_value(SCRIPT_LICENSE_COLUMN, f"{repo.get_license_id()}")

    #STARS
    entry.set_cell_value(SCRIPT_STARS_COLUMN, f"{repo.stargazers_count}")

    #KLOC
    if hasattr(repo.stat, "loc"):
        entry.set_cell_value(SCRIPT_KLOC_COLUMN, f"{(repo.stat.loc)/1000:.0f}")
    else:
        logging.warning(f"repo.stat.loc is None for {entry.url}")

    #contributors total
    fill_in_contributors_stat(entry.get_vendor(), repo, entry)

    #main language
    entry.set_cell_value(SCRIPT_LANGUAGE_COLUMN, repo.language)

    #active
    entry.set_cell_value(SCRIPT_PROJECT_ACTIVE_SCORE_COLUMN, repo.get_active_score())
    
    #files
    f1, f2, f3 = repo.get_top3_files()
    entry.set_cell_value(SCRIPT_TOP3_FILES_COLUMN, f"{f1}, {f2}, {f3}")

    #size
    entry.set_cell_value(SCRIPT_SIZE_COLUMN, f"{repo.size/1024:.0f}")

    #forks
    entry.set_cell_value(SCRIPT_FORKS_COLUMN, repo.forks_count)

    #issues
    #entry.set_cell_value(SCRIPT_ISSUES_COLUMN, repo.open_issues_count)

    #description
    entry.set_cell_value(SCRIPT_PROJECT_DESCRIPTION_COLUMN, repo.description)

    workbook.save('modified_oss_projects.xlsx')
    logging.debug(f"Updated values for {entry.sheet.title} row {entry.row_index} {entry.url}")

    
def fill_in_contributors_stat(vendor, repo, entry):
    contributors_data = generate_contributors_report([vendor], repo)

    total_contributors = contributors_data["n_total_contributors"]
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_TOTAL_COLUMN, total_contributors)

    total_users = contributors_data["total_users"]
    total_users_percent = f"{total_users/total_contributors}" if total_contributors > 0 else "0"
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_USERS_TOTAL_COLUMN, total_users)
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_USERS_TOTAL_PERCENT_COLUMN, total_users_percent)

    total_anon = contributors_data["total_anon"]
    total_anon_percent = f"{total_anon/total_contributors}" if total_contributors > 0 else "0"
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_ANON_TOTAL_COLUMN, total_anon)
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_ANON_TOTAL_PERCENT_COLUMN, total_anon_percent)

    n_total_contributors_matched_users = contributors_data["n_total_contributors_matched_users"]
    n_total_contributors_matched_users_p = f"{n_total_contributors_matched_users/total_contributors}" if total_contributors > 0 else "0"
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN, n_total_contributors_matched_users)
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN, n_total_contributors_matched_users_p)

    n_total_contributors_others = total_contributors - n_total_contributors_matched_users
    n_total_contributors_others_p = f"{n_total_contributors_others/total_contributors}" if total_contributors > 0 else "0"
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_COLUMN, n_total_contributors_others)
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN, n_total_contributors_others_p)    

    
    total_contributions = contributors_data["total_contributions"]
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_TOTAL_CONTRIBUTIONS_COLUMN, total_contributions)

    top3_contributors = f"{contributors_data["top1_contributor"], contributors_data["top2_contributor"], contributors_data["top3_contributor"]}"
    entry.set_cell_value(SCRIPT_CONTRIBUTORS_TOP3_COLUMN, top3_contributors)

    
    n_total_active_users = contributors_data["n_total_active_users"]
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_COLUMN, n_total_active_users)
    active_contributors_percent = n_total_active_users/total_contributors if total_contributors > 0 else 0
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_PERCENT_COLUMN, active_contributors_percent)

    n_total_active_users_matched = contributors_data["n_total_active_users_matched"]
    n_total_active_users_matched_p = n_total_active_users_matched/n_total_active_users if n_total_active_users > 0 else 0
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_COLUMN, n_total_active_users_matched)
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_INTERNAL_P_COLUMN, n_total_active_users_matched_p)

    n_total_active_users_others = n_total_active_users - n_total_active_users_matched
    n_total_active_users_others_p = n_total_active_users_others/n_total_active_users if n_total_active_users > 0 else 0
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_COLUMN, n_total_active_users_others)
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_TOTAL_OTHERS_P_COLUMN, n_total_active_users_others_p)

    
    authors_in_contributors = contributors_data["authors_in_contributors"]
    authors_in_contributors_percent = authors_in_contributors/n_total_active_users if n_total_active_users > 0 else 0
    authors_in_users = contributors_data["authors_in_users"]
    authors_in_users_percent = authors_in_users/n_total_active_users if n_total_active_users > 0 else 0
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_COLUMN, authors_in_contributors)
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_PRESENTED_ON_GITHUB_P_COLUMN, authors_in_contributors_percent)
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_COLUMN, authors_in_users)
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_GITHUB_USERS_P_COLUMN, authors_in_users_percent)

    #domains
    d1, d2, d3 = repo.get_top3_domains()
    entry.set_cell_value(SCRIPT_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN, f"{d1}, {d2}, {d3}")

    top_users_domains = contributors_data["top_users_domains"]
    entry.set_cell_value(SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_COLUMN, top_users_domains)
    top_users_companies = contributors_data["top_users_companies"]
    entry.set_cell_value(SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_COLUMN, top_users_companies)
    top_companies_with_organization = contributors_data["top_companies_with_organization"]
    entry.set_cell_value(SCRIPT_TOP10_CONTRIBUTORS_COMPANIES_WITHIN_ORG_COLUMN, top_companies_with_organization)
    top_domains_with_organization = contributors_data["top_domains_with_organization"]
    entry.set_cell_value(SCRIPT_TOP10_CONTRIBUTORS_DOMAINS_WITH_ORG_COLUMN, top_domains_with_organization)
    top_active_users_domains = contributors_data["top_active_users_domains"]
    entry.set_cell_value(SCRIPT_TOP10_ACTIVE_CONTRIBUTORS_DOMAINS_COLUMN, top_active_users_domains)
